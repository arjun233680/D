#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Build the classification worklist from already-extracted workbooks.

Classifying Q91-150 is the only part of this pipeline that needs judgement, and
it does not need the PDFs at all — just the question text. This writes one small
CSV per paper holding the question number, the English stem and an empty
`subject` column, plus the allowed category list for that paper as a comment.

A session then fills in the `subject` column and nothing else. Each file is a
few kilobytes, so a session can hold several papers at once instead of
re-parsing 700 KB PDFs one at a time.

    python scripts/make_labels.py
    -> labels/todo/htet-2024-tgt-5201-hindi.csv   (60 rows, subject blank)

When a file is filled in, move it to labels/ and to_import_format.py picks it
up with --labels-dir labels/.
"""
from __future__ import annotations

import argparse
import csv
import glob
import os
import re

from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
OUT = os.path.join(ROOT, "output")
LABELS = os.path.join(ROOT, "labels")
TODO = os.path.join(LABELS, "todo")
CATEGORIES_MD = os.path.join(ROOT, "papers", "CATEGORIES.md")

# Q1-90 subjects that follow from position alone and need no judgement
FIXED = "1-30=CDP,31-45=Hindi,46-60=English"


def categories_for(code: str) -> list[str]:
    """Pull the allowed category list for a subject code out of CATEGORIES.md."""
    if not os.path.exists(CATEGORIES_MD):
        return []
    for line in open(CATEGORIES_MD, encoding="utf-8"):
        if not line.startswith("|"):
            continue
        cells = [c.strip() for c in line.strip().strip("|").split("|")]
        if len(cells) >= 3 and cells[0] == code:
            return [c.strip() for c in re.split(r"·|\u00b7", cells[2]) if c.strip()]
    return []


def main(argv=None) -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--only", help="filter by level or subject code")
    args = ap.parse_args(argv)

    os.makedirs(TODO, exist_ok=True)
    made = 0

    for path in sorted(glob.glob(os.path.join(OUT, "htet-*.xlsx"))):
        stem = os.path.splitext(os.path.basename(path))[0]
        if stem.endswith("-import"):
            continue
        m = re.match(r"htet-(\d{4})-(prt|tgt|pgt)-(\d{4})-(.+)", stem)
        if not m:
            continue
        year, level, code, subject = m.groups()
        if args.only and args.only not in (level, code):
            continue

        done = os.path.join(LABELS, f"{stem}.csv")
        if os.path.exists(done):
            print(f"  skip (already labelled)  {stem}")
            continue

        ws = load_workbook(path, data_only=True)["questions"]
        header = [str(c.value) for c in ws[1]]
        qi = header.index("q_no")
        eni = header.index("question_english")
        hii = header.index("question_hindi")

        rows = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            q = r[qi]
            if not isinstance(q, int):
                continue
            if q <= 60:              # position rules cover these
                continue
            stem_text = (r[eni] or r[hii] or "")
            rows.append((q, re.sub(r"\s+", " ", str(stem_text))[:300]))

        cats = categories_for(code)
        dest = os.path.join(TODO, f"{stem}.csv")
        with open(dest, "w", newline="", encoding="utf-8-sig") as fh:
            w = csv.writer(fh)
            w.writerow([f"# {stem} — fill the subject column only"])
            w.writerow([f"# allowed: {' | '.join(cats) if cats else 'SEE papers/CATEGORIES.md'}"])
            w.writerow(["# Q61-90 use: Numerical Aptitude | Reasoning | Haryana GK | General GK"])
            w.writerow(["# anything that fits none of them: REVIEW"])
            w.writerow(["q_no", "subject", "question"])
            for q, text in rows:
                w.writerow([q, "", text])
        made += 1
        print(f"  {len(rows):3} rows  {dest}")

    print(f"\n{made} worklist file(s) in {TODO}")
    print("Fill the `subject` column, then move the file to labels/ and run:")
    print(f"  python scripts/to_import_format.py output/<paper>.xlsx \\")
    print(f"      --labels-dir labels/ --fixed \"{FIXED}\" --categories \"...\"")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
