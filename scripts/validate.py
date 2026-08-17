#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Validate an extracted HTET workbook before it is allowed to become output.

Checks, all of which must pass:
  * expected row count (150 for a full paper, 60 for an elective-only run)
  * every answer matches the official key, including dropped and multi answers
  * four options present for every question
  * the language-blank pattern: Q31-45 Hindi only, Q46-60 English only,
    everything else bilingual (only applies to a full 150-row run)
  * no duplicate question text within the file

Exits non-zero on failure so a pipeline can stop instead of writing a bad file.

    python scripts/validate.py output/htet-2024-tgt-5211.xlsx \
        --key papers/htet-2024-tgt-5211-science-key.pdf --expect 150
"""
from __future__ import annotations

import argparse
import re
import sys

from openpyxl import load_workbook

from parse_key import parse_key

HI_ONLY = range(31, 46)
EN_ONLY = range(46, 61)


def norm(s) -> str:
    return re.sub(r"\s+", "", str(s or ""))


def main(argv=None) -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("workbook")
    ap.add_argument("--key", required=True, help="official answer key PDF")
    ap.add_argument("--expect", type=int, default=None,
                    help="expected row count (150 or 60)")
    ap.add_argument("--sheet", default="questions")
    args = ap.parse_args(argv)

    wb = load_workbook(args.workbook, data_only=True)
    ws = wb[args.sheet] if args.sheet in wb.sheetnames else wb[wb.sheetnames[0]]
    header = [str(c.value).strip() if c.value else "" for c in ws[1]]
    idx = {h: i for i, h in enumerate(header)}

    rows = [r for r in ws.iter_rows(min_row=2, values_only=True)
            if any(v not in (None, "") for v in r)]
    key = parse_key(args.key)
    errors: list[str] = []

    expect = args.expect or len(rows)
    if len(rows) != expect:
        errors.append(f"row count is {len(rows)}, expected {expect}")

    seen: dict[str, int] = {}
    for r in rows:
        q = r[idx["q_no"]]
        try:
            q = int(q)
        except (TypeError, ValueError):
            errors.append(f"bad q_no {q!r}")
            continue

        want_ans, want_type = key.get(q, (None, None))
        got_ans = r[idx["answer"]] if "answer" in idx else None
        got_type = r[idx["answer_type"]] if "answer_type" in idx else None
        got_ans = "" if got_ans in (None, "") else str(got_ans).replace(" ", "")
        want_ans = "" if want_ans is None else want_ans
        if got_ans != want_ans or got_type != want_type:
            errors.append(
                f"Q{q}: answer {got_ans!r}/{got_type!r}, key says {want_ans!r}/{want_type!r}")

        hi = norm(r[idx["question_hindi"]]) if "question_hindi" in idx else ""
        en = norm(r[idx["question_english"]]) if "question_english" in idx else ""
        if expect == 150:
            if q in HI_ONLY and en:
                errors.append(f"Q{q}: English should be blank (Hindi language section)")
            elif q in EN_ONLY and hi:
                errors.append(f"Q{q}: Hindi should be blank (English language section)")
            elif q not in HI_ONLY and q not in EN_ONLY and not (hi and en):
                errors.append(f"Q{q}: expected both languages")
        else:
            if not (hi and en):
                errors.append(f"Q{q}: expected both languages")

        for lang in ("hindi", "english"):
            if q in HI_ONLY and lang == "english":
                continue
            if q in EN_ONLY and lang == "hindi":
                continue
            for n in range(1, 5):
                col = f"option_{n}_{lang}"
                if col in idx and not norm(r[idx[col]]):
                    errors.append(f"Q{q}: {col} is empty")

        fp = hi or en
        if fp:
            if fp in seen:
                errors.append(f"Q{q}: duplicate of Q{seen[fp]}")
            seen[fp] = q

    if errors:
        print(f"FAIL {args.workbook} — {len(errors)} problem(s):", file=sys.stderr)
        for e in errors[:40]:
            print(f"  {e}", file=sys.stderr)
        if len(errors) > 40:
            print(f"  ... and {len(errors) - 40} more", file=sys.stderr)
        return 1

    print(f"OK {args.workbook}: {len(rows)} rows, all answers match the official key")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
