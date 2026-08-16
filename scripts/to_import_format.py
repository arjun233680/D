#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Convert an HTET extraction workbook into the app's import format.

Input: the extraction workbook, whose "questions" sheet carries
    q_no, part,
    question_hindi,   option_1_hindi   .. option_4_hindi,
    question_english, option_1_english .. option_4_english,
    answer, answer_type, ai_note

Output: the import layout
    Exam, Level, Subject, Topic, Question EN, Question HI,
    Option A EN, Option B EN, Option C EN, Option D EN,
    Option A HI, Option B HI, Option C HI, Option D HI,
    Correct Answer, Explanation EN, Explanation HI, Year, Paper,
    answer_type, ai_note

Topic and both Explanation columns are deliberately left blank.
answer_type and ai_note are carried over unchanged as trailing extras.

Correct Answer
    1 / 2 / 3 / 4    ->  A / B / C / D
    "3,4"            ->  "C,D"     (answer_type = multiple)
    blank            ->  blank     (answer_type = dropped)

The Subject column is filled from two independent sources:

  --fixed   position rules for the ranges whose subject follows purely from
            where the question sits, e.g. "1-30=CDP,31-45=Hindi,46-60=English"

  --labels  a CSV/JSON of q_no -> subject for every question whose subject can
            only be decided by reading the question. This file is produced by
            reading the paper; the script never guesses a subject and never
            derives one from the paper's printed blueprint, which is often
            wrong about which block is which.

Every value in the labels file must appear in --categories (or be the literal
REVIEW). A question covered by neither --fixed nor --labels becomes REVIEW and
is listed in the report, so nothing is silently mislabelled.

Examples
    python scripts/to_import_format.py output/htet-2024-tgt-5211.xlsx \\
        --level TGT --paper Science --year 2024 \\
        --fixed "1-30=CDP,31-45=Hindi,46-60=English" \\
        --labels labels/htet-2024-tgt-5211.csv \\
        --categories "Numerical Aptitude,Reasoning,Haryana GK,General GK,\\
Physics,Chemistry,Biology,Science Pedagogy"

    python scripts/to_import_format.py output/*.xlsx -o import/ --labels-dir labels/
"""
from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sys
from collections import Counter

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------
# output layout
# --------------------------------------------------------------------------

IMPORT_COLUMNS = [
    "Exam", "Level", "Subject", "Topic",
    "Question EN", "Question HI",
    "Option A EN", "Option B EN", "Option C EN", "Option D EN",
    "Option A HI", "Option B HI", "Option C HI", "Option D HI",
    "Correct Answer", "Explanation EN", "Explanation HI",
    "Year", "Paper",
]
EXTRA_COLUMNS = ["answer_type", "ai_note"]
ALL_COLUMNS = IMPORT_COLUMNS + EXTRA_COLUMNS

REVIEW = "REVIEW"
LETTER = {"1": "A", "2": "B", "3": "C", "4": "D"}

SOURCE_MAP = {
    "question_english": "Question EN",
    "question_hindi": "Question HI",
    "option_1_english": "Option A EN",
    "option_2_english": "Option B EN",
    "option_3_english": "Option C EN",
    "option_4_english": "Option D EN",
    "option_1_hindi": "Option A HI",
    "option_2_hindi": "Option B HI",
    "option_3_hindi": "Option C HI",
    "option_4_hindi": "Option D HI",
}

LEVELS = {"prt": "PRT", "tgt": "TGT", "pgt": "PGT"}
PAPER_OVERRIDES = {"sst": "Social Studies", "gk": "GK"}

DEFAULT_CATEGORIES = [
    "Numerical Aptitude", "Reasoning", "Haryana GK", "General GK",
    "Physics", "Chemistry", "Biology", "Science Pedagogy",
]


# --------------------------------------------------------------------------
# arguments that describe the paper
# --------------------------------------------------------------------------

def infer_meta(path: str) -> dict:
    """Best-effort Exam / Level / Year / Paper from `htet-2024-tgt-5211-science`.

    `Paper` is the subject name from the filename ("science" -> "Science"), not
    the numeric subject code. Anything not found stays None; --flags override.
    """
    stem = os.path.splitext(os.path.basename(path))[0]
    tokens = [t for t in re.split(r"[-_\s]+", stem.lower()) if t]
    meta = {"Exam": None, "Level": None, "Year": None, "Paper": None}
    words: list[str] = []

    for tok in tokens:
        if tok in LEVELS and meta["Level"] is None:
            meta["Level"] = LEVELS[tok]
        elif re.fullmatch(r"(19|20)\d{2}", tok) and meta["Year"] is None:
            meta["Year"] = int(tok)
        elif tok.isdigit():
            continue                       # subject code — not used as Paper
        elif meta["Exam"] is None and tok.isalpha() and len(tok) <= 6 and not words:
            meta["Exam"] = tok.upper()
        elif tok.isalpha():
            words.append(tok)

    if words:
        slug = "-".join(words)
        meta["Paper"] = PAPER_OVERRIDES.get(
            slug, " ".join(w.capitalize() for w in words))
    return meta


def parse_fixed(spec: str) -> list[tuple[int, int, str]]:
    """"1-30=CDP,31-45=Hindi,46-60=English" -> [(1,30,'CDP'), ...]"""
    rules = []
    for chunk in (spec or "").split(","):
        chunk = chunk.strip()
        if not chunk:
            continue
        if "=" not in chunk:
            raise SystemExit(f"--fixed: '{chunk}' is not RANGE=SUBJECT")
        rng, subject = chunk.split("=", 1)
        subject = subject.strip()
        rng = rng.strip()
        m = re.fullmatch(r"(\d+)\s*(?:-\s*(\d+))?", rng)
        if not m or not subject:
            raise SystemExit(f"--fixed: cannot read '{chunk}'")
        lo = int(m.group(1))
        hi = int(m.group(2) or m.group(1))
        if hi < lo:
            raise SystemExit(f"--fixed: range '{rng}' runs backwards")
        rules.append((lo, hi, subject))

    rules.sort()
    for (a_lo, a_hi, a_s), (b_lo, b_hi, b_s) in zip(rules, rules[1:]):
        if b_lo <= a_hi:
            raise SystemExit(
                f"--fixed: ranges {a_lo}-{a_hi}={a_s} and {b_lo}-{b_hi}={b_s} overlap")
    return rules


def load_labels(path: str, allowed: set, warnings: list) -> dict:
    """Read q_no -> subject from a CSV (q_no,subject) or a JSON object."""
    if not path:
        return {}
    if not os.path.exists(path):
        raise SystemExit(f"--labels: {path} not found")

    labels: dict[int, str] = {}
    if path.lower().endswith(".json"):
        raw = json.load(open(path, encoding="utf-8"))
        raw = raw.get("labels", raw) if isinstance(raw, dict) else raw
        items = raw.items()
    else:
        with open(path, newline="", encoding="utf-8-sig") as fh:
            rows = list(csv.reader(fh))
        if rows and not rows[0][0].strip().isdigit():
            rows = rows[1:]                      # header
        items = [(r[0], r[1]) for r in rows if len(r) >= 2 and r[0].strip()]

    for q, subject in items:
        try:
            qno = int(str(q).strip())
        except ValueError:
            warnings.append(f"{path}: question number {q!r} is not a number")
            continue
        subject = str(subject).strip()
        if not subject:
            continue
        if subject != REVIEW and subject not in allowed:
            warnings.append(
                f"{path}: Q{qno} has subject {subject!r}, which is not in --categories")
            subject = REVIEW
        labels[qno] = subject
    return labels


# --------------------------------------------------------------------------
# conversion
# --------------------------------------------------------------------------

def convert_answer(raw, answer_type: str, where: str, warnings: list) -> str:
    """Map the official answer onto letters. Never invents an answer."""
    if answer_type == "dropped":
        if raw not in (None, ""):
            warnings.append(f"{where}: answer_type=dropped but answer is {raw!r}")
        return ""
    if raw is None or str(raw).strip() == "":
        if answer_type != "dropped":
            warnings.append(f"{where}: blank answer but answer_type={answer_type!r}")
        return ""

    parts = [p.strip().rstrip(".") for p in re.split(r"[,&/]| and ", str(raw)) if p.strip()]
    letters = []
    for p in parts:
        if p not in LETTER:
            warnings.append(f"{where}: cannot map answer part {p!r} (from {raw!r})")
            return ""
        letters.append(LETTER[p])
    if len(letters) > 1 and answer_type != "multiple":
        warnings.append(f"{where}: {len(letters)} answers but answer_type={answer_type!r}")
    return ",".join(letters)


def subject_for(qno: int, fixed, labels: dict) -> str:
    for lo, hi, subject in fixed:
        if lo <= qno <= hi:
            return subject
    return labels.get(qno, REVIEW)


def read_source(path: str, sheet: str | None, fixed, labels, warnings) -> list[dict]:
    wb = load_workbook(path, data_only=True)
    ws = None
    if sheet:
        if sheet not in wb.sheetnames:
            raise SystemExit(f"{path}: no sheet named {sheet!r}")
        ws = wb[sheet]
    else:
        ws = wb["questions"] if "questions" in wb.sheetnames else wb[wb.sheetnames[0]]

    header = [str(c.value).strip() if c.value is not None else ""
              for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(header)}
    missing = [h for h in list(SOURCE_MAP) + ["answer"] if h not in idx]
    if missing:
        raise SystemExit(f"{path}: sheet {ws.title!r} is missing column(s): "
                         + ", ".join(missing))

    rows = []
    for i, raw in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        if all(v is None or str(v).strip() == "" for v in raw):
            continue

        def get(col):
            j = idx.get(col)
            if j is None or j >= len(raw) or raw[j] is None:
                return ""
            return str(raw[j]).strip()

        qtext = get("q_no")
        qno = int(qtext) if qtext.isdigit() else i
        atype = get("answer_type")

        rec = {c: "" for c in ALL_COLUMNS}
        for src, dst in SOURCE_MAP.items():
            rec[dst] = get(src)
        rec["Subject"] = subject_for(qno, fixed, labels)
        rec["Correct Answer"] = convert_answer(
            raw[idx["answer"]] if idx["answer"] < len(raw) else None,
            atype, f"{os.path.basename(path)} Q{qno}", warnings)
        rec["answer_type"] = atype
        rec["ai_note"] = get("ai_note")
        rec["_q_no"] = qno
        rows.append(rec)
    return rows


# --------------------------------------------------------------------------
# writing
# --------------------------------------------------------------------------

def write_xlsx(rows: list[dict], path: str, sheet_name: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    hdr_fill = PatternFill("solid", fgColor="1F3864")
    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    body_font = Font(name="Arial", size=10)
    review_fill = PatternFill("solid", fgColor="FCE4D6")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append(ALL_COLUMNS)
    for c in ws[1]:
        c.fill, c.font, c.border = hdr_fill, hdr_font, border
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    subj_col = ALL_COLUMNS.index("Subject") + 1
    for r in rows:
        ws.append([r[c] for c in ALL_COLUMNS])
        for c in ws[ws.max_row]:
            c.font, c.border = body_font, border
            c.alignment = Alignment(vertical="top", wrap_text=True)
        if r["Subject"] == REVIEW:
            ws.cell(ws.max_row, subj_col).fill = review_fill

    widths = {"Exam": 8, "Level": 8, "Subject": 20, "Topic": 14,
              "Question EN": 46, "Question HI": 46, "Correct Answer": 14,
              "Explanation EN": 16, "Explanation HI": 16,
              "Year": 8, "Paper": 14, "answer_type": 12, "ai_note": 60}
    for i, name in enumerate(ALL_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(name, 26)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(ALL_COLUMNS))}{ws.max_row}"

    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    wb.save(path)


def write_csv(rows: list[dict], path: str) -> None:
    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8-sig") as fh:
        w = csv.DictWriter(fh, fieldnames=ALL_COLUMNS, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def report(dest: str, meta: dict, rows: list[dict], fixed) -> None:
    counts = Counter(r["Subject"] for r in rows)
    review = [r["_q_no"] for r in rows if r["Subject"] == REVIEW]
    fixed_subjects = {s for _, _, s in fixed}

    print(f"\n{dest}")
    print(f"  {len(rows)} rows | {meta['Exam']} / {meta['Level']} / "
          f"{meta['Paper']} / {meta['Year']}")
    print("  Subject counts:")
    for name, n in sorted(counts.items(), key=lambda kv: (-kv[1], kv[0])):
        src = "position" if name in fixed_subjects else ("--" if name == REVIEW else "content")
        print(f"    {name:<20} {n:>4}   ({src})")
    answered = sum(1 for r in rows if r["Correct Answer"])
    multi = sum(1 for r in rows if "," in r["Correct Answer"])
    print(f"  Correct Answer: {answered} filled ({multi} multi), "
          f"{len(rows) - answered} blank")
    if review:
        print(f"  REVIEW ({len(review)}): "
              + ", ".join(f"Q{q}" for q in review))
    else:
        print("  REVIEW: none")


# --------------------------------------------------------------------------
# cli
# --------------------------------------------------------------------------

def main(argv=None) -> int:
    ap = argparse.ArgumentParser(
        description="Convert HTET extraction workbooks to the app import format.",
        formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("inputs", nargs="+", help="extraction .xlsx file(s)")
    ap.add_argument("-o", "--out",
                    help="output file (single input) or directory (multiple). "
                         "Default: next to the input, '-import' appended.")
    ap.add_argument("--exam", default="HTET")
    ap.add_argument("--level", help="e.g. TGT. Inferred from the filename if omitted.")
    ap.add_argument("--paper", help="e.g. Science. Inferred from the filename if omitted.")
    ap.add_argument("--year", help="e.g. 2024. Inferred from the filename if omitted.")
    ap.add_argument("--fixed", default="",
                    help='position-based Subject rules, e.g. '
                         '"1-30=CDP,31-45=Hindi,46-60=English"')
    ap.add_argument("--labels",
                    help="CSV (q_no,subject) or JSON of content-based Subject values")
    ap.add_argument("--labels-dir",
                    help="directory of <input-stem>.csv label files, for batch runs")
    ap.add_argument("--categories", default=",".join(DEFAULT_CATEGORIES),
                    help="comma-separated list of Subject values the labels file "
                         "may use; anything else becomes REVIEW")
    ap.add_argument("--sheet", help="source sheet name (default: 'questions')")
    ap.add_argument("--sheet-name", default="questions",
                    help="sheet name in the output workbook")
    ap.add_argument("--csv", action="store_true", help="also write a UTF-8 CSV")
    args = ap.parse_args(argv)

    many = len(args.inputs) > 1
    if many and args.out and os.path.splitext(args.out)[1]:
        ap.error("--out must be a directory when several inputs are given")
    if many and args.labels:
        ap.error("use --labels-dir, not --labels, with several inputs")

    allowed = {c.strip() for c in args.categories.split(",") if c.strip()}
    fixed = parse_fixed(args.fixed)
    for _, _, subject in fixed:
        allowed.add(subject)

    warnings: list[str] = []
    failed = 0

    for src in args.inputs:
        if not os.path.exists(src):
            print(f"!! {src}: not found", file=sys.stderr)
            failed += 1
            continue

        stem = os.path.splitext(os.path.basename(src))[0]
        meta = infer_meta(src)
        meta["Exam"] = args.exam or meta["Exam"]
        if args.level:
            meta["Level"] = args.level
        if args.paper:
            meta["Paper"] = args.paper
        if args.year:
            meta["Year"] = int(args.year)

        label_path = args.labels
        if not label_path and args.labels_dir:
            for ext in (".csv", ".json"):
                cand = os.path.join(args.labels_dir, stem + ext)
                if os.path.exists(cand):
                    label_path = cand
                    break
        labels = load_labels(label_path, allowed, warnings)

        rows = read_source(src, args.sheet, fixed, labels, warnings)
        for r in rows:
            for k in ("Exam", "Level", "Year", "Paper"):
                r[k] = meta[k] if meta[k] is not None else ""

        if args.out and not many and os.path.splitext(args.out)[1]:
            dest = args.out
        else:
            folder = args.out or os.path.dirname(os.path.abspath(src))
            dest = os.path.join(folder, f"{stem}-import.xlsx")

        write_xlsx(rows, dest, args.sheet_name)
        if args.csv:
            write_csv(rows, os.path.splitext(dest)[0] + ".csv")

        report(dest, meta, rows, fixed)
        blank = [k for k, v in meta.items() if v in (None, "")]
        if blank:
            print(f"  note: {', '.join(blank)} not set — pass --{blank[0].lower()}")
        if not label_path:
            print("  note: no labels file used, so every question outside --fixed "
                  "is REVIEW")

    if warnings:
        print()
        for w in warnings:
            print(f"!! {w}", file=sys.stderr)
    return 1 if failed or warnings else 0


if __name__ == "__main__":
    raise SystemExit(main())
