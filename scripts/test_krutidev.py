#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Regression tests for the Kruti Dev converter.

Two layers:
  1. WORDS  — hand-checked word pairs. All of these pass today.
  2. GOLDEN — tests/golden_5211.json holds the 150 hand-verified rows of
     htet-2024-tgt-5211-science. Point the extractor at that paper and compare;
     this is the target the converter has to reach before a bulk run.

    python scripts/test_krutidev.py
"""
import json
import os
import re
import sys

from krutidev import convert

WORDS = [
    ("Hkkx", "भाग"),
    ("cky fodkl", "बाल विकास"),
    ("f'k{kk'kkL=", "शिक्षाशास्त्र"),
    ("fuEufyf[kr", "निम्नलिखित"),
    ("fØ;k", "क्रिया"),
    ("vkjEHk", "आरम्भ"),
    ("djus", "करने"),
    ("izfØ;k", "प्रक्रिया"),
    ("deZ", "कर्म"),
    ("dk;Z", "कार्य"),
    ("iwoZ", "पूर्व"),
    ("vfHkçsj.kk", "अभिप्रेरणा"),
    ("iz'uksa", "प्रश्नों"),
    ("mÙkj", "उत्तर"),
    ("vfHko`fÙk", "अभिवृत्ति"),
    ("dkSu&lh", "कौन-सी"),
    ("f'k{kk", "शिक्षा"),
]


def run_words() -> int:
    bad = 0
    for src, want in WORDS:
        got = convert(src, digits=False)
        if got != want:
            bad += 1
            print(f"  FAIL {src!r} -> {got!r}, want {want!r}")
    print(f"WORDS: {len(WORDS) - bad}/{len(WORDS)} pass")
    return bad


def load_golden():
    path = os.path.join(os.path.dirname(__file__), "tests", "golden_5211.json")
    if not os.path.exists(path):
        print("GOLDEN: tests/golden_5211.json missing — skipped")
        return None
    return json.load(open(path, encoding="utf-8"))["5211"]


def run_golden(extracted_path=None) -> int:
    """Compare an extraction of paper 5211 against the hand-verified rows."""
    golden = load_golden()
    if golden is None:
        return 0
    if not extracted_path:
        print(f"GOLDEN: {len(golden)} reference rows loaded. "
              "Pass an extracted .xlsx to compare against them.")
        return 0

    from openpyxl import load_workbook
    ws = load_workbook(extracted_path, data_only=True)["questions"]
    got = {r[0]: r for r in ws.iter_rows(min_row=2, values_only=True)}
    n = bad = 0
    for row in golden:
        q = row["q_no"]
        if q not in got:
            print(f"  MISSING Q{q}")
            bad += 1
            continue
        pairs = [(row["question_hindi"], got[q][2]), (row["question_english"], got[q][7])]
        pairs += list(zip(row["options_hindi"], got[q][3:7]))
        pairs += list(zip(row["options_english"], got[q][8:12]))
        for want, mine in pairs:
            if not want:
                continue
            n += 1
            if re.sub(r"\s+", "", str(want)) != re.sub(r"\s+", "", str(mine or "")):
                bad += 1
                if bad <= 15:
                    print(f"  Q{q}: got {str(mine)[:60]!r}\n       want {str(want)[:60]!r}")
    print(f"GOLDEN: {n - bad}/{n} cells match ({(n - bad) / max(n, 1) * 100:.1f}%)")
    return bad


if __name__ == "__main__":
    failures = run_words()
    failures += run_golden(sys.argv[1] if len(sys.argv) > 1 else None)
    raise SystemExit(1 if failures else 0)
