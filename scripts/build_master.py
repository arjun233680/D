#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Merge every extracted paper + its subject labels into one master workbook.

Input: output/htet-*.xlsx (from extract_all.py, one row per question) and
labels/htet-*.csv (from make_labels.py, q_no -> subject for Q61-150).

Each paper's own workbook already holds exactly the questions it
contributes to the deduplicated 2310-question set -- the two common-section
source papers (TGT 5211, PGT 5316) carry the full Q1-150, PRT 5101 carries
its own full Q1-150, and every other paper carries only its elective
Q91-150 (see papers/EXTRACTION-RULES.md section 3). This script does not
re-derive or duplicate the common section; it just reads what is already on
disk once per paper.

Output: output/HTET-2024-MASTER.xlsx
    Sheet 1 "Questions"     every row, in Exam/Level/Paper/q_no order
    Sheet 2 "Needs Review"  rows with status=CHECK, with a reason
    Sheet 3 "Summary"       one line per paper + overall subject counts

    python scripts/build_master.py
"""
from __future__ import annotations

import glob
import os
import re
from collections import Counter

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
OUT = os.path.join(ROOT, "output")
LABELS = os.path.join(ROOT, "labels")
DEST = os.path.join(OUT, "HTET-2024-MASTER.xlsx")

LEVEL_NAMES = {"prt": "PRT", "tgt": "TGT", "pgt": "PGT"}
REVIEW = "REVIEW"
LETTER = {"1": "A", "2": "B", "3": "C", "4": "D"}

# Q1-60 subject follows from position alone, on the three papers that carry
# it (see EXTRACTION-RULES.md section 2/3) -- every other paper's workbook
# only has Q91-150 rows to begin with, so this never applies to them.
FIXED = [(1, 30, "CDP"), (31, 45, "Hindi"), (46, 60, "English")]

QUESTIONS_COLUMNS = [
    "Exam", "Level", "Subject", "Topic",
    "Question EN", "Question HI",
    "Option A EN", "Option B EN", "Option C EN", "Option D EN",
    "Option A HI", "Option B HI", "Option C HI", "Option D HI",
    "Correct Answer", "Explanation EN", "Explanation HI",
    "Year", "Paper", "answer_type", "q_no", "status",
]


def paper_name(subject_slug: str) -> str:
    """"social-studies" -> "Social Studies" """
    return " ".join(w.capitalize() for w in subject_slug.split("-"))


def load_labels(stem: str) -> dict[int, str]:
    path = os.path.join(LABELS, f"{stem}.csv")
    labels: dict[int, str] = {}
    if not os.path.exists(path):
        return labels
    with open(path, encoding="utf-8-sig") as fh:
        lines = [ln for ln in fh if not ln.lstrip().startswith("#")]
    import csv
    rows = list(csv.reader(lines))
    for row in rows[1:]:                      # skip the q_no,subject,question header
        if len(row) < 2 or not row[0].strip().isdigit():
            continue
        subject = row[1].strip()
        if subject:
            labels[int(row[0])] = subject
    return labels


def subject_for(qno: int, labels: dict[int, str]) -> str:
    for lo, hi, subject in FIXED:
        if lo <= qno <= hi:
            return subject
    return labels.get(qno, REVIEW)


def convert_answer(raw, answer_type: str) -> str:
    if raw is None or str(raw).strip() == "":
        return ""
    parts = [p.strip().rstrip(".") for p in re.split(r"[,&/]| and ", str(raw)) if p.strip()]
    letters = [LETTER[p] for p in parts if p in LETTER]
    if len(letters) != len(parts):
        return ""                              # unrecognised token -- never guess
    return ",".join(letters)


def status_and_reason(subject: str, answer_type: str, correct: str) -> tuple[str, str]:
    if subject == REVIEW:
        return "CHECK", "subject needs manual classification"
    if answer_type == "dropped":
        return "OK", ""
    if not correct:
        return "CHECK", "no official answer parsed"
    return "OK", ""


def rows_for_paper(path: str) -> list[dict]:
    stem = os.path.splitext(os.path.basename(path))[0]
    m = re.match(r"htet-(\d{4})-(prt|tgt|pgt)-(\d{4})-(.+)", stem)
    if not m:
        print(f"!! skipping unrecognised file name: {stem}")
        return []
    year, level, code, subject_slug = m.groups()
    labels = load_labels(stem)

    wb = load_workbook(path, data_only=True)
    ws = wb["questions"] if "questions" in wb.sheetnames else wb.active
    header = [str(c.value) for c in ws[1]]
    idx = {h: i for i, h in enumerate(header)}

    rows = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        qno = raw[idx["q_no"]]
        if not isinstance(qno, int):
            continue

        def get(col):
            j = idx.get(col)
            v = raw[j] if j is not None and j < len(raw) else None
            return str(v).strip() if v is not None else ""

        subject = subject_for(qno, labels)
        answer_type = get("answer_type")
        correct = convert_answer(raw[idx["answer"]] if idx.get("answer") is not None else None,
                                  answer_type)
        status, reason = status_and_reason(subject, answer_type, correct)

        rows.append({
            "Exam": "HTET", "Level": LEVEL_NAMES[level], "Subject": subject, "Topic": "",
            "Question EN": get("question_english"), "Question HI": get("question_hindi"),
            "Option A EN": get("option_1_english"), "Option B EN": get("option_2_english"),
            "Option C EN": get("option_3_english"), "Option D EN": get("option_4_english"),
            "Option A HI": get("option_1_hindi"), "Option B HI": get("option_2_hindi"),
            "Option C HI": get("option_3_hindi"), "Option D HI": get("option_4_hindi"),
            "Correct Answer": correct, "Explanation EN": "", "Explanation HI": "",
            "Year": int(year), "Paper": paper_name(subject_slug),
            "answer_type": answer_type, "q_no": qno, "status": status,
            "_reason": reason, "_stem": stem,
        })
    return rows


def style_header(ws, ncols: int) -> None:
    fill = PatternFill("solid", fgColor="1F3864")
    font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    for c in ws[1]:
        c.fill, c.font = fill, font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    if ws.max_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{ws.max_row}"


def write_questions_sheet(wb: Workbook, rows: list[dict]) -> None:
    ws = wb.active
    ws.title = "Questions"
    ws.append(QUESTIONS_COLUMNS)
    body_font = Font(name="Arial", size=10)
    review_fill = PatternFill("solid", fgColor="FCE4D6")
    for r in rows:
        ws.append([r[c] for c in QUESTIONS_COLUMNS])
        for c in ws[ws.max_row]:
            c.font = body_font
            c.alignment = Alignment(vertical="top", wrap_text=True)
        if r["status"] == "CHECK":
            ws.cell(ws.max_row, QUESTIONS_COLUMNS.index("status") + 1).fill = review_fill
    widths = {"Question EN": 44, "Question HI": 44, "Subject": 22, "Paper": 16,
              "Option A EN": 22, "Option B EN": 22, "Option C EN": 22, "Option D EN": 22,
              "Option A HI": 22, "Option B HI": 22, "Option C HI": 22, "Option D HI": 22}
    for i, name in enumerate(QUESTIONS_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(name, 12)
    style_header(ws, len(QUESTIONS_COLUMNS))


def write_review_sheet(wb: Workbook, rows: list[dict]) -> None:
    ws = wb.create_sheet("Needs Review")
    cols = ["Exam", "Level", "Paper", "q_no", "Subject", "status", "Reason",
            "Question EN", "Question HI"]
    ws.append(cols)
    body_font = Font(name="Arial", size=10)
    for r in rows:
        if r["status"] != "CHECK":
            continue
        ws.append([r["Exam"], r["Level"], r["Paper"], r["q_no"], r["Subject"],
                   r["status"], r["_reason"], r["Question EN"], r["Question HI"]])
        for c in ws[ws.max_row]:
            c.font = body_font
            c.alignment = Alignment(vertical="top", wrap_text=True)
    for i, name in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = \
            44 if "Question" in name else (10 if name in ("q_no", "status") else 18)
    style_header(ws, len(cols))


def write_summary_sheet(wb: Workbook, rows: list[dict]) -> None:
    ws = wb.create_sheet("Summary")
    by_paper: dict[str, list[dict]] = {}
    for r in rows:
        by_paper.setdefault(r["_stem"], []).append(r)

    cols = ["Level", "Paper", "Rows", "OK", "CHECK"]
    ws.append(cols)
    for stem in sorted(by_paper):
        prs = by_paper[stem]
        ok = sum(1 for r in prs if r["status"] == "OK")
        check = sum(1 for r in prs if r["status"] == "CHECK")
        ws.append([prs[0]["Level"], f"{prs[0]['Paper']} ({stem.split('-')[3]})",
                   len(prs), ok, check])
    for i, w in enumerate((10, 34, 8, 8, 8), start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    style_header(ws, len(cols))

    ws.append([])
    start = ws.max_row + 1
    ws.append(["Subject", "Count"])
    for c in ws[start]:
        c.font = Font(name="Arial", bold=True, size=10)
    counts = Counter(r["Subject"] for r in rows)
    for subject, n in sorted(counts.items(), key=lambda kv: (-kv[1], kv[0])):
        ws.append([subject, n])


def main() -> int:
    paper_paths = sorted(
        p for p in glob.glob(os.path.join(OUT, "htet-*.xlsx"))
        if not os.path.basename(p).startswith("HTET-"))

    rows: list[dict] = []
    for path in paper_paths:
        rows.extend(rows_for_paper(path))
    rows.sort(key=lambda r: (r["Level"], r["_stem"], r["q_no"]))

    wb = Workbook()
    write_questions_sheet(wb, rows)
    write_review_sheet(wb, rows)
    write_summary_sheet(wb, rows)
    os.makedirs(OUT, exist_ok=True)
    wb.save(DEST)

    total = len(rows)
    ok = sum(1 for r in rows if r["status"] == "OK")
    check = total - ok
    review = sum(1 for r in rows if r["Subject"] == REVIEW)
    print(f"{DEST}: {total} rows from {len(paper_paths)} papers")
    print(f"  OK: {ok}  CHECK: {check}  (of which subject=REVIEW: {review})")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
